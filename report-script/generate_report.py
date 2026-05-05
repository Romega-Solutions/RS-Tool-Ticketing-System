"""
RS Weekly Report Generator
Pulls task data from Plane.so API and generates the 7-section weekly report as .xlsx.

Sections auto-populated:
  1. Header (name + coverage dates)
  4. Pending Projects (active tasks)
  5. Key Accomplishments (tasks completed this week)

Sections left blank for IC input:
  2. Client Engagement Activities
  3. Risks / Issues / Roadblocks
  6. Ideas / Recommendations
  7. Management Remarks

Usage:
  python generate_report.py                     # Current week, all users
  python generate_report.py --user "Ken Garcia" # Current week, one user
  python generate_report.py --week 2026-05-05   # Specific week (Monday date)
  python generate_report.py --bulk              # All users in one workbook
"""

import argparse
import json
import os
import sys
import time
from datetime import datetime, timedelta

import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

load_dotenv()

# --- Configuration ---

PLANE_BASE_URL = os.getenv("PLANE_BASE_URL", "").rstrip("/")
PLANE_API_KEY = os.getenv("PLANE_API_KEY", "")
PLANE_WORKSPACE_SLUG = os.getenv("PLANE_WORKSPACE_SLUG", "romega")
PLANE_PROJECT_SLUG = os.getenv("PLANE_PROJECT_SLUG", "")
OUTPUT_DIR = os.getenv("REPORT_OUTPUT_DIR", "./reports")

# Status mapping: Plane status → report display value
STATUS_MAP = {
    "backlog": "Drafted",
    "unstarted": "Drafted",
    "started": "On-going",
    "todo": "Drafted",
    "in_progress": "On-going",
    "in_review": "For Approval",
    "completed": "Completed",
}

# Statuses that count as "pending" (shown in Section 4)
PENDING_STATUSES = {
    "backlog",
    "unstarted",
    "started",
    "todo",
    "in_progress",
    "in_review",
}

# --- Styling ---

RS_BLUE = "0069D9"
RS_ORANGE = "D97B00"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F5F7"

HEADER_FONT = Font(name="Merriweather", size=14, bold=True, color=RS_BLUE)
SECTION_FONT = Font(name="Source Sans 3", size=12, bold=True, color=WHITE)
SECTION_FILL_BLUE = PatternFill(start_color=RS_BLUE, end_color=RS_BLUE, fill_type="solid")
SECTION_FILL_ORANGE = PatternFill(start_color=RS_ORANGE, end_color=RS_ORANGE, fill_type="solid")
COL_HEADER_FONT = Font(name="Source Sans 3", size=10, bold=True, color="333333")
COL_HEADER_FILL = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
BODY_FONT = Font(name="Source Sans 3", size=10, color="333333")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")


# --- Plane API Client ---


class PlaneClient:
    """Minimal client for Plane.so REST API."""

    def __init__(self, base_url: str, api_key: str, workspace_slug: str):
        if not base_url or not api_key:
            raise ValueError(
                "PLANE_BASE_URL and PLANE_API_KEY must be set. "
                "See .env.example for details."
            )
        self.base_url = base_url
        self.workspace_slug = workspace_slug
        self.session = requests.Session()
        self.session.headers.update({
            "X-API-Key": api_key,
            "Content-Type": "application/json",
        })

    def _get(self, path: str, params: dict | None = None, _retries: int = 3) -> dict:
        url = f"{self.base_url}/api/v1/workspaces/{self.workspace_slug}{path}"
        for attempt in range(_retries):
            resp = self.session.get(url, params=params, timeout=30)
            if resp.status_code == 429:
                wait = 2 ** attempt
                print(f"[rate limit] waiting {wait}s...", flush=True)
                time.sleep(wait)
                continue
            resp.raise_for_status()
            return resp.json()
        raise RuntimeError(f"Rate limit: failed after {_retries} retries: {url}")

    def get_projects(self) -> list[dict]:
        data = self._get("/projects/")
        return data.get("results", data) if isinstance(data, dict) else data

    def get_members(self) -> list[dict]:
        data = self._get("/members/")
        return data.get("results", data) if isinstance(data, dict) else data

    def get_issues(self, project_id: str, params: dict | None = None) -> list[dict]:
        # /work-items/ is the current endpoint; fall back to /issues/ for older Plane instances.
        for endpoint in (
            f"/projects/{project_id}/work-items/",
            f"/projects/{project_id}/issues/",
        ):
            try:
                all_results: list[dict] = []
                cursor = None
                while True:
                    page_params = {**(params or {}), "per_page": 100}
                    if cursor:
                        page_params["cursor"] = cursor
                    data = self._get(endpoint, params=page_params)
                    results = data.get("results", []) if isinstance(data, dict) else (data or [])
                    all_results.extend(results)
                    next_cursor = data.get("next_cursor") if isinstance(data, dict) else None
                    has_more = data.get("next_page_results", False) if isinstance(data, dict) else False
                    if not has_more or not next_cursor:
                        break
                    cursor = next_cursor
                return all_results
            except requests.exceptions.HTTPError as exc:
                if exc.response is not None and exc.response.status_code == 404:
                    continue
                raise
        raise RuntimeError(f"Neither /work-items/ nor /issues/ endpoint available for project {project_id}")

    def get_project_states(self, project_id: str) -> list[dict]:
        data = self._get(f"/projects/{project_id}/states/")
        return data.get("results", data) if isinstance(data, dict) else data

    def get_project_name(self, projects: list[dict], project_id: str) -> str:
        for p in projects:
            if p.get("id") == project_id:
                return p.get("name", "Unknown Project")
        return "Unknown Project"


# --- Week Calculation ---


def get_week_range(week_start_str: str | None = None) -> tuple[datetime, datetime]:
    """Return (Monday 00:00, Friday 23:59) for the given or current week."""
    if week_start_str:
        monday = datetime.strptime(week_start_str, "%Y-%m-%d")
    else:
        today = datetime.now()
        monday = today - timedelta(days=today.weekday())
    monday = monday.replace(hour=0, minute=0, second=0, microsecond=0)
    friday = monday + timedelta(days=4, hours=23, minutes=59, seconds=59)
    return monday, friday


# --- Data Fetching ---


def normalize_status(value: str) -> str:
    return str(value).strip().lower().replace(" ", "_")


def build_state_group_lookup(states: list[dict]) -> dict[str, str]:
    """Map Plane state UUID -> normalized status group for a project."""
    lookup: dict[str, str] = {}
    for state in states:
        state_id = state.get("id")
        if not state_id:
            continue
        group = state.get("group") or state.get("key") or state.get("name")
        if group:
            lookup[str(state_id)] = normalize_status(group)
    return lookup


def extract_assignee_ids(issue: dict) -> set[str] | None:
    """
    Return assignee IDs when determinable.
    - set(...) when explicit assignee fields exist
    - None when fields are absent/unknown
    """
    has_assignee_fields = "assignees" in issue or "assignee_ids" in issue
    if not has_assignee_fields:
        return None

    result: set[str] = set()
    assignees = issue.get("assignees")
    assignee_ids = issue.get("assignee_ids")

    if isinstance(assignees, list):
        for item in assignees:
            if isinstance(item, str):
                result.add(item)
            elif isinstance(item, dict) and item.get("id"):
                result.add(str(item.get("id")))

    if isinstance(assignee_ids, list):
        for item in assignee_ids:
            if item:
                result.add(str(item))

    return result


def extract_status_group(issue: dict, state_group_lookup: dict[str, str] | None = None) -> str:
    """Normalize Plane issue state/group across dict and string payload shapes."""
    state_detail = issue.get("state_detail")
    state_group = issue.get("state_group")
    state = issue.get("state")

    if isinstance(state_detail, dict):
        value = state_detail.get("group")
        if value:
            return normalize_status(value)

    if state_group:
        return normalize_status(state_group)

    if isinstance(state, dict):
        state_id = state.get("id")
        if state_id and state_group_lookup:
            mapped = state_group_lookup.get(str(state_id))
            if mapped:
                return mapped

        for key in ("group", "key", "name"):
            value = state.get(key)
            if value:
                return normalize_status(value)

    if isinstance(state, str):
        if state_group_lookup:
            mapped = state_group_lookup.get(state)
            if mapped:
                return mapped
        return normalize_status(state)

    return ""


def fetch_user_tasks(
    client: PlaneClient,
    projects: list[dict],
    user_id: str,
    debug: bool = False,
) -> tuple[list[dict], list[dict]]:
    """Fetch pending tasks and completed tasks for a user across all projects."""
    pending = []
    completed = []

    for project in projects:
        pid = project["id"]
        pname = project.get("name", "Unknown")
        states = client.get_project_states(pid)
        state_group_lookup = build_state_group_lookup(states)

        issues = client.get_issues(pid, params={"assignee": user_id})

        for issue in issues:
            issue_assignees = extract_assignee_ids(issue)
            if issue_assignees is not None and user_id not in issue_assignees:
                continue

            status_group = extract_status_group(issue, state_group_lookup=state_group_lookup)
            completed_at = issue.get("completed_at")

            if debug:
                print(json.dumps({
                    "project": pname,
                    "name": issue.get("name"),
                    "state": issue.get("state"),
                    "state_detail": issue.get("state_detail"),
                    "state_group": issue.get("state_group"),
                    "assignees": issue.get("assignees"),
                    "assignee_ids": issue.get("assignee_ids"),
                    "target_date": issue.get("target_date"),
                    "completed_at": issue.get("completed_at"),
                    "updated_at": issue.get("updated_at"),
                    "_state_group_lookup_hit": state_group_lookup.get(str(issue.get("state"))),
                    "_issue_assignee_ids": sorted(issue_assignees) if issue_assignees is not None else None,
                    "_normalized_status_group": status_group,
                }, indent=2, ensure_ascii=False))

            issue["_project_name"] = pname
            issue["_status_group"] = status_group

            if status_group == "completed" or completed_at:
                completed.append(issue)
            elif status_group in {"cancelled", "canceled", "archived"}:
                continue
            else:
                # Keep unknown non-terminal states visible in Pending Projects.
                pending.append(issue)

    return pending, completed


def filter_completed_this_week(
    tasks: list[dict], week_start: datetime, week_end: datetime
) -> list[dict]:
    """Filter completed tasks to only those completed within the given week."""
    result = []
    for task in tasks:
        completed_at = task.get("completed_at")
        if not completed_at:
            continue
        try:
            dt = datetime.fromisoformat(completed_at.replace("Z", "+00:00")).replace(tzinfo=None)
            if week_start <= dt <= week_end:
                result.append(task)
        except (ValueError, TypeError):
            continue
    return result


def extract_member_user(member: dict) -> dict:
    """Normalize member payload across API shapes: {member:{}}, {user:{}}, or flat."""
    if not isinstance(member, dict):
        return {}
    return member.get("member") or member.get("user") or member


def member_matches_query(member: dict, query: str) -> bool:
    """Match user query against id, display name, username, name, or email."""
    user_info = extract_member_user(member)
    q = query.lower()
    candidates = [
        user_info.get("id"),
        user_info.get("display_name"),
        user_info.get("username"),
        user_info.get("name"),
        user_info.get("email"),
    ]
    for value in candidates:
        if value and q in str(value).lower():
            return True
    return False


# --- Excel Generation ---


def write_section_header(ws, row: int, text: str, fill=None) -> int:
    """Write a section header row with merged cells. Returns next row."""
    if fill is None:
        fill = SECTION_FILL_BLUE
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = SECTION_FONT
    cell.fill = fill
    cell.alignment = Alignment(vertical="center")
    return row + 1


def write_column_headers(ws, row: int, headers: list[str]) -> int:
    """Write column header row. Returns next row."""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = COL_HEADER_FONT
        cell.fill = COL_HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = WRAP_ALIGNMENT
    return row + 1


def write_data_row(ws, row: int, values: list, col_count: int = 5) -> int:
    """Write a data row. Returns next row."""
    for col_idx in range(1, col_count + 1):
        cell = ws.cell(
            row=row,
            column=col_idx,
            value=values[col_idx - 1] if col_idx - 1 < len(values) else "",
        )
        cell.font = BODY_FONT
        cell.border = THIN_BORDER
        cell.alignment = WRAP_ALIGNMENT
    return row + 1


def write_empty_row(ws, row: int, text: str = "(No items)", col_count: int = 5) -> int:
    """Write a placeholder row when a section has no data."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name="Source Sans 3", size=10, italic=True, color="999999")
    cell.alignment = Alignment(horizontal="center")
    return row + 1


def write_blank_section(ws, row: int, text: str = "(Fill in manually)") -> int:
    """Write a blank area for manual IC input."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 2, end_column=5)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name="Source Sans 3", size=10, italic=True, color="999999")
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    for r in range(row, row + 3):
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = THIN_BORDER
    return row + 3


def generate_ic_sheet(
    ws,
    user_name: str,
    week_start: datetime,
    week_end: datetime,
    pending_tasks: list[dict],
    completed_tasks: list[dict],
):
    """Generate a single IC's weekly report sheet."""
    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 25

    row = 1

    # --- Row 1: Title ---
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    cell = ws.cell(row=row, column=1, value="ROMEGA SOLUTIONS — WEEKLY REPORT")
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center")
    row += 1

    # --- Row 2: Name + Coverage ---
    ws.cell(row=row, column=1, value="Name:").font = COL_HEADER_FONT
    ws.cell(row=row, column=2, value=user_name).font = BODY_FONT
    ws.cell(row=row, column=3, value="Coverage:").font = COL_HEADER_FONT
    coverage = f"{week_start.strftime('%B %d')} – {week_end.strftime('%B %d, %Y')}"
    ws.cell(row=row, column=4, value=coverage).font = BODY_FONT
    row += 2  # blank row

    # --- Section 2: Client Engagement Activities ---
    row = write_section_header(ws, row, "CLIENT ENGAGEMENT ACTIVITIES")
    row = write_column_headers(ws, row, ["Activity", "Date", "Details", "", ""])
    row = write_blank_section(ws, row)
    row += 1

    # --- Section 3: Risks / Issues / Roadblocks ---
    row = write_section_header(ws, row, "RISKS / ISSUES / ROADBLOCKS", SECTION_FILL_ORANGE)
    row = write_column_headers(ws, row, ["Description", "Resolution", "Escalation?", "", ""])
    row = write_blank_section(ws, row)
    row += 1

    # --- Section 4: Pending Projects (AUTO-POPULATED) ---
    row = write_section_header(ws, row, "PENDING PROJECTS (auto-populated from tasks)")
    row = write_column_headers(
        ws, row, ["Project Name", "Task Title", "TAT Estimate", "Status", "Remarks"]
    )

    if pending_tasks:
        for task in pending_tasks:
            project_name = task.get("_project_name", "")
            title = task.get("name", "Untitled")
            status_group = task.get("_status_group", "")
            display_status = STATUS_MAP.get(status_group, status_group.replace("_", " ").title())

            # TAT estimate
            due_date_str = task.get("target_date")
            if due_date_str:
                try:
                    due = datetime.strptime(due_date_str, "%Y-%m-%d")
                    days_left = (due - datetime.now()).days
                    if days_left < 0:
                        tat = f"Overdue by {abs(days_left)} days"
                    elif days_left == 0:
                        tat = "Due today"
                    else:
                        tat = f"{days_left} days remaining"
                except ValueError:
                    tat = "No deadline set"
            else:
                tat = "No deadline set"

            # Remarks from blocker description or label
            remarks = ""
            labels = task.get("label_detail", []) or task.get("labels", [])
            if isinstance(labels, list):
                for label in labels:
                    name = label.get("name", "") if isinstance(label, dict) else str(label)
                    if "blocker" in name.lower():
                        remarks = task.get("description_stripped", "")[:100] or "Blocker"
                        break

            row = write_data_row(ws, row, [project_name, title, tat, display_status, remarks])
    else:
        row = write_empty_row(ws, row)

    row += 1

    # --- Section 5: Key Accomplishments (AUTO-POPULATED) ---
    row = write_section_header(ws, row, "KEY ACCOMPLISHMENTS (auto-populated from tasks)")
    row = write_column_headers(ws, row, ["Description", "Completion Date", "Remarks", "", ""])

    if completed_tasks:
        for task in completed_tasks:
            title = task.get("name", "Untitled")
            completed_at = task.get("completed_at") or task.get("updated_at", "")
            try:
                dt = datetime.fromisoformat(completed_at.replace("Z", "+00:00")).replace(tzinfo=None)
                date_str = dt.strftime("%B %d, %Y")
            except (ValueError, TypeError):
                date_str = ""
            row = write_data_row(ws, row, [title, date_str, ""])
    else:
        row = write_empty_row(ws, row)

    row += 1

    # --- Section 6: Ideas / Recommendations ---
    row = write_section_header(ws, row, "IDEAS / RECOMMENDATIONS")
    row = write_blank_section(ws, row)
    row += 1

    # --- Section 7: Management Remarks ---
    row = write_section_header(ws, row, "MANAGEMENT REMARKS", SECTION_FILL_ORANGE)
    row = write_blank_section(ws, row, "(Manager fills after submission)")


def sanitize_sheet_name(name: str) -> str:
    """Create a valid Excel sheet name from a user's display name."""
    parts = name.strip().split()
    if len(parts) >= 2:
        sheet = f"{parts[-1]}, {parts[0][0]}"
    else:
        sheet = name
    # Excel sheet names max 31 chars, no special chars
    for ch in r"[]:*?/\\":
        sheet = sheet.replace(ch, "")
    return sheet[:31]


# --- Main ---


def main():
    parser = argparse.ArgumentParser(
        description="Generate RS Weekly Report from Plane.so",
        allow_abbrev=False,
    )
    parser.add_argument("--week", help="Monday date of the week (YYYY-MM-DD). Default: current week.")
    parser.add_argument("--user", help="Generate report for a specific user (name, username, email, or id). Default: all users.")
    parser.add_argument("--bulk", action="store_true", help="Generate one workbook with all users (one sheet per IC).")
    parser.add_argument("--dry-run", action="store_true", help="Print what would be generated without calling the API.")
    parser.add_argument("--debug-members", action="store_true", help="Print member matching fields and raw member payload.")
    parser.add_argument("--debug-issues", action="store_true", help="Print Plane issue payload/status fields.")
    args = parser.parse_args()

    week_start, week_end = get_week_range(args.week)
    print(f"Report period: {week_start.strftime('%B %d')} – {week_end.strftime('%B %d, %Y')}")

    if args.dry_run:
        print("[DRY RUN] Would connect to Plane API and generate reports.")
        print(f"  Base URL: {PLANE_BASE_URL}")
        print(f"  Workspace: {PLANE_WORKSPACE_SLUG}")
        print(f"  Output dir: {OUTPUT_DIR}")
        return

    client = PlaneClient(PLANE_BASE_URL, PLANE_API_KEY, PLANE_WORKSPACE_SLUG)
    projects = client.get_projects()

    # Filter by project slug/identifier if configured
    if PLANE_PROJECT_SLUG:
        projects = [
            p for p in projects
            if p.get("identifier") == PLANE_PROJECT_SLUG or p.get("slug") == PLANE_PROJECT_SLUG
        ]
        if not projects:
            print(f"No project found matching '{PLANE_PROJECT_SLUG}'")
            sys.exit(1)
        print(f"Filtering to project: {projects[0].get('name')} ({PLANE_PROJECT_SLUG})")

    members = client.get_members()

    print(f"Found {len(projects)} projects, {len(members)} members")

    if args.debug_members:
        print("\n[DEBUG] Member candidates:")
        for idx, member in enumerate(members, start=1):
            user_info = extract_member_user(member)
            print(
                f"  {idx}. id={user_info.get('id', '-')}, "
                f"display_name={user_info.get('display_name', '-')}, "
                f"username={user_info.get('username', '-')}, "
                f"name={user_info.get('name', '-')}, "
                f"email={user_info.get('email', '-')}"
            )
        print("\n[DEBUG] Raw member payload:")
        print(json.dumps(members, indent=2, ensure_ascii=False))

    # Filter to specific user if requested
    if args.user:
        members = [
            m for m in members
            if member_matches_query(m, args.user)
        ]
        if not members:
            print(f"No member found matching '{args.user}'")
            sys.exit(1)

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    week_label = week_start.strftime("%Y-%m-%d")

    if args.bulk:
        # One workbook, one sheet per IC
        wb = Workbook()
        wb.remove(wb.active)

        for member in members:
            user_info = extract_member_user(member)
            user_id = user_info.get("id", "")
            user_name = user_info.get("display_name") or user_info.get("name") or "Unknown"

            print(f"  Generating sheet for {user_name}...")
            pending, completed = fetch_user_tasks(client, projects, user_id, debug=args.debug_issues)
            completed = filter_completed_this_week(completed, week_start, week_end)

            sheet_name = sanitize_sheet_name(user_name)
            ws = wb.create_sheet(title=sheet_name)
            generate_ic_sheet(ws, user_name, week_start, week_end, pending, completed)

        filename = f"{week_label} - RS Weekly Report (All).xlsx"
        filepath = os.path.join(OUTPUT_DIR, filename)
        wb.save(filepath)
        print(f"\nBulk report saved: {filepath}")

    else:
        # Individual workbooks per user
        for member in members:
            user_info = extract_member_user(member)
            user_id = user_info.get("id", "")
            user_name = user_info.get("display_name") or user_info.get("name") or "Unknown"

            print(f"  Generating report for {user_name}...")
            pending, completed = fetch_user_tasks(client, projects, user_id, debug=args.debug_issues)
            completed = filter_completed_this_week(completed, week_start, week_end)

            wb = Workbook()
            ws = wb.active
            ws.title = "Weekly Report"
            generate_ic_sheet(ws, user_name, week_start, week_end, pending, completed)

            safe_name = user_name.replace(" ", "_")
            filename = f"{week_label} - {safe_name} - Weekly Report.xlsx"
            filepath = os.path.join(OUTPUT_DIR, filename)
            wb.save(filepath)
            print(f"  Saved: {filepath}")

    print("\nDone.")


if __name__ == "__main__":
    main()
